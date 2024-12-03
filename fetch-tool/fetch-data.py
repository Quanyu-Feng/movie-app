import openpyxl
import requests

API_KEY = 'a7c925c73a6cb8fbaca41e94ff265c4b'
EXCEL_FILE_PATH = 'assets/data/movies.xlsx'

def fetch_genre_mapping():
    url = f"https://api.themoviedb.org/3/genre/movie/list?api_key={API_KEY}&language=zh-CN"
    response = requests.get(url)
    response.raise_for_status()
    data = response.json()
    return {genre['id']: genre['name'] for genre in data['genres']}

def fetch_movie_data(movie_name):
    url = f"https://api.themoviedb.org/3/search/movie?api_key={API_KEY}&query={movie_name}&language=zh-CN"
    response = requests.get(url)
    response.raise_for_status()
    data = response.json()
    if data['results']:
        movie_id = data['results'][0]['id']
        # Fetch detailed movie data
        details_url = f"https://api.themoviedb.org/3/movie/{movie_id}?api_key={API_KEY}&append_to_response=credits,releases&language=zh-CN"
        details_response = requests.get(details_url)
        details_response.raise_for_status()
        return details_response.json()
    return None

def update_excel_with_movie_data():
    genre_mapping = fetch_genre_mapping()
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        movie_name = row[0]
        if movie_name:
            movie_data = fetch_movie_data(movie_name)
            if movie_data:
                # Correctly map genre_ids to genre names
                genre_names = [genre_mapping.get(genre['id'], "Unknown") for genre in movie_data['genres']]
                genre_names_str = ', '.join(genre_names)

                # Extract additional details
                backdrop_path = movie_data.get('backdrop_path', '')
                runtime = movie_data.get('runtime', '')
                vote_average = movie_data.get('vote_average', '')
                overview = movie_data.get('overview', '')

                # Extract certification
                certification = ''
                if 'releases' in movie_data:
                    for country in movie_data['releases']['countries']:
                        if country['iso_3166_1'] == 'US':  # Assuming US certification
                            certification = country.get('certification', '')
                            break

                # Extract cast and directors
                cast = ', '.join([member['name'] for member in movie_data['credits']['cast'][:10]])
                directors = ', '.join([member['name'] for member in movie_data['credits']['crew'] if member['job'] == 'Director'])

                # Store data in the Excel file with the new sequence
                sheet.cell(row=row_index, column=1, value=movie_data['title'])  # Title
                sheet.cell(row=row_index, column=2, value=genre_names_str)      # Genre
                sheet.cell(row=row_index, column=3, value=movie_data['release_date'])  # Release Date
                sheet.cell(row=row_index, column=4, value=movie_data.get('poster_path', ''))  # Poster Path
                sheet.cell(row=row_index, column=5, value=vote_average)         # Vote Average
                sheet.cell(row=row_index, column=6, value=movie_data['id'])     # ID
                sheet.cell(row=row_index, column=7, value=backdrop_path)        # Backdrop Path
                sheet.cell(row=row_index, column=8, value=runtime)              # Runtime
                sheet.cell(row=row_index, column=9, value=vote_average)         # Vote Average (again)
                sheet.cell(row=row_index, column=10, value=overview)            # Overview
                sheet.cell(row=row_index, column=11, value=certification)       # Certification
                sheet.cell(row=row_index, column=12, value=cast)                # Cast
                sheet.cell(row=row_index, column=13, value=directors)           # Directors

    workbook.save(EXCEL_FILE_PATH)

if __name__ == "__main__":
    update_excel_with_movie_data()
